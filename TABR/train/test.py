from collections import deque
import numpy as np
import os
from tqdm import tqdm

import torch
import torch.nn.functional as F
import math
from model import Actor, Critic
import env as env_valid
import fixed_env as env_test
import load_trace
import openpyxl

RANDOM_SEED = 2024
S_INFO = 6
S_LEN = 8
A_DIM = 4

TACTILE_BIT_RATE = [[2.51976, 3.59517, 4.86501, 8.08464],
                    [6.71976, 7.79517, 9.06501, 11.28464],
                    [9.38751, 10.20676, 11.2304, 14.91361],
                    [12.39386, 13.23207, 14.26618, 16.92485]]

TEST_TRACES_VALID = '../dataset/network/lvs_trace_final/fixed/'
SUMMARY_DIR = './Results/sim/'
LOG_FILE = './Results/sim//log'
TEST_LOG_FOLDER = './Results/sim/test_results/'
MILLISECONDS_IN_SECOND = 1000.0
LOG_FILE_VALID = './Results/sim/test_results/log_valid_ppo'
TEST_LOG_FOLDER_VALID = './Results/sim/test_results/'
DEFAULT_BIT_RATE_LEVEL = 0
DEFAULT_DELAY_LEVEL = 0
Log_path = './Results/sim'
TACTILE_CHUNCK_LEN_64 = 64 / 2800
LATENCY_LIMIT = 45 / MILLISECONDS_IN_SECOND
REBUF_PENALTY = 1.85
SKIP_PENALTY = 1.85
SMOOTH_PENALTY = 0.2
LANTENCY_PENALTY = 10

dtype = torch.cuda.FloatTensor if torch.cuda.is_available() else torch.FloatTensor
dlongtype = torch.cuda.LongTensor if torch.cuda.is_available() else torch.LongTensor
dshorttype = torch.cuda.ShortTensor if torch.cuda.is_available() else torch.ShortTensor

wb = openpyxl.Workbook()
del wb['Sheet']

np.set_printoptions(formatter={'float': '{: 0.3f}'.format})


def evaluation(model, log_path_ini, net_env, all_file_name, detail_log=True):
    state = np.zeros((S_INFO, S_LEN))
    state = torch.from_numpy(state)

    last_a_bit_rate = DEFAULT_BIT_RATE_LEVEL
    last_a_delay = DEFAULT_DELAY_LEVEL

    log_path = log_path_ini + '_' + all_file_name[net_env.trace_idx]
    log_file = open(log_path, 'w')
    sheets = []

    for i in range(1, 6):
        sheet = wb.create_sheet(title=f"Sheet{i}")
        sheets.append(sheet)

    time_stamp = 0
    for tactile_count in tqdm(range(len(all_file_name))):
        a_bit_rate = last_a_bit_rate
        a_delay = last_a_delay
        reward_all = 0
        reward_all_A = 0
        reward_all_B = 0
        send_data_size_all = 0
        rebuf_all = 0
        buffer_size_all = 0
        end_delay_all = 0
        skip_frame_time_len_all = 0
        i = 0
        download_all = 0
        count = 0
        if not detail_log:
            sheet_line_index = 0
            while True:

                time, time_interval, send_data_size, chunk_len, \
                    rebuf, buffer_size, play_time_len, end_delay, \
                    cdn_newest_id, download_id, cdn_has_frame, skip_frame_time_len, decision_flag, \
                    buffer_flag, cdn_flag, skip_flag, end_of_tactile = \
                    net_env.get_tactile_frame(a_bit_rate, a_delay)

                rebuf_all += rebuf
                skip_frame_time_len_all += skip_frame_time_len

                if end_delay >= LATENCY_LIMIT:
                    MEET_DEAD_LINE = 0
                else:
                    MEET_DEAD_LINE = 1

                if not cdn_flag:
                    count = count + 1
                    send_data_size_all += send_data_size
                    if TACTILE_BIT_RATE[a_delay][a_bit_rate] <= TACTILE_BIT_RATE[2][0]:
                        bit_rate = TACTILE_BIT_RATE[a_delay][a_bit_rate]
                    else:
                        bit_rate = TACTILE_BIT_RATE[2][0] + 2 * math.log10(
                            TACTILE_BIT_RATE[a_delay][a_bit_rate] / TACTILE_BIT_RATE[2][0])

                    reward_frame = (
                            bit_rate * play_time_len
                            * MEET_DEAD_LINE
                            - REBUF_PENALTY * rebuf
                            - SKIP_PENALTY * skip_frame_time_len)

                    reward_frame_A = bit_rate * play_time_len * MEET_DEAD_LINE
                    reward_frame_B = REBUF_PENALTY * rebuf + SKIP_PENALTY * skip_frame_time_len

                    buffer_size_all += buffer_size
                    end_delay_all += end_delay
                    download_all += time_interval
                    i = i + 1
                else:
                    if TACTILE_BIT_RATE[a_delay][a_bit_rate] <= TACTILE_BIT_RATE[2][0]:
                        bit_rate = TACTILE_BIT_RATE[a_delay][a_bit_rate]
                    else:
                        bit_rate = TACTILE_BIT_RATE[2][0] + 2 * math.log10(
                            TACTILE_BIT_RATE[a_delay][a_bit_rate] / TACTILE_BIT_RATE[2][0])

                    reward_frame = (bit_rate * play_time_len
                                    * MEET_DEAD_LINE
                                    - REBUF_PENALTY * rebuf
                                    - SKIP_PENALTY * skip_frame_time_len)
                    reward_frame_A = bit_rate * play_time_len * MEET_DEAD_LINE
                    reward_frame_B = REBUF_PENALTY * rebuf + SKIP_PENALTY * skip_frame_time_len

                if decision_flag or end_of_tactile:
                    # exit()
                    if TACTILE_BIT_RATE[a_delay][a_bit_rate] <= TACTILE_BIT_RATE[2][0]:
                        bit_rate = TACTILE_BIT_RATE[a_delay][a_bit_rate]
                    else:
                        bit_rate = TACTILE_BIT_RATE[2][0] + 2 * math.log10(
                            TACTILE_BIT_RATE[a_delay][a_bit_rate] / TACTILE_BIT_RATE[2][0])

                    if TACTILE_BIT_RATE[last_a_delay][last_a_bit_rate] <= TACTILE_BIT_RATE[2][0]:
                        last_bit_rate = TACTILE_BIT_RATE[last_a_delay][last_a_bit_rate]
                    else:
                        last_bit_rate = TACTILE_BIT_RATE[2][0] + 2 * math.log10(
                            TACTILE_BIT_RATE[last_a_delay][last_a_bit_rate] / TACTILE_BIT_RATE[2][0])

                    reward_all += -1 * SMOOTH_PENALTY * (
                        abs(bit_rate - last_bit_rate))

                    reward_all += - LANTENCY_PENALTY * end_delay_all / count
                    reward_all_C = LANTENCY_PENALTY * end_delay_all / count
                    reward_all_D = (SMOOTH_PENALTY * (abs(bit_rate - last_bit_rate)))

                    last_a_bit_rate = a_bit_rate
                    last_a_delay = a_delay

                    state = np.roll(state, -1, axis=1)
                    state[0, -1] = TACTILE_BIT_RATE[a_delay][a_bit_rate] / float(TACTILE_BIT_RATE[-1][-1])
                    state[1, -1] = float(end_delay_all / count / LATENCY_LIMIT)
                    state[2, -1] = float(send_data_size_all / count /
                                         TACTILE_BIT_RATE[-1][-1] * TACTILE_CHUNCK_LEN_64)
                    state[3, -1] = float(download_all / count / LATENCY_LIMIT)
                    state[4, -1] = float(buffer_size_all / count / LATENCY_LIMIT)
                    state[5, -1] = float((skip_frame_time_len_all + rebuf_all) / count / LATENCY_LIMIT)

                    state = torch.from_numpy(state)
                    buffer_size_all = 0

                    with torch.no_grad():
                        prob_bit, prob_delay = model(state.unsqueeze(0).type(dtype))

                    action_bit = prob_bit.multinomial(num_samples=1).detach()
                    action_delay = prob_delay.multinomial(num_samples=1).detach()

                    log_file.write(str(time_stamp) + '\t' +
                                   str(TACTILE_BIT_RATE[a_delay][a_bit_rate]) + '\t' +
                                   str(buffer_size) + '\t' +
                                   str(rebuf) + '\t' +
                                   str(send_data_size) + '\t' +
                                   str(end_delay) + '\t' +
                                   str(reward_all) + '\n')
                    log_file.flush()

                    a_bit_rate = int(action_bit.squeeze().cpu().numpy())
                    a_delay = int(action_delay.squeeze().cpu().numpy())

                    inner_values = [reward_all, reward_all_A, reward_all_B, reward_all_C, reward_all_D]

                    sheet_line_index = sheet_line_index + 1
                    for sheet_idx, sheet in enumerate(sheets):
                        column = tactile_count + 1
                        row = sheet_line_index + 1
                        sheet.cell(row=row, column=column, value=inner_values[sheet_idx])

                    reward_all = 0
                    reward_all_A = 0
                    reward_all_B = 0
                    download_all = 0
                    reward_frame_A = 0
                    reward_frame_B = 0
                    reward_frame = 0
                    end_delay_all = 0
                    rebuf_all = 0
                    send_data_size_all = 0
                    skip_frame_time_len_all = 0
                    count = 0

                if end_of_tactile:
                    # exit()
                    state = np.zeros((S_INFO, S_LEN))
                    state = torch.from_numpy(state)
                    last_a_bit_rate = DEFAULT_BIT_RATE_LEVEL
                    last_a_delay = DEFAULT_DELAY_LEVEL
                    if tactile_count + 1 >= len(all_file_name):
                        break
                    else:
                        log_path = log_path_ini + '_' + all_file_name[net_env.trace_idx]
                        log_file = open(log_path, 'w')
                        break

                reward_all += reward_frame
                reward_all_A += reward_frame_A
                # print(reward_all_A)
                reward_all_B += reward_frame_B

    wb.save("output_orig.xlsx")


def valid(shared_model, epoch, log_file):
    os.system('rm -r ' + TEST_LOG_FOLDER_VALID)
    os.system('mkdir ' + TEST_LOG_FOLDER_VALID)

    model = Actor(A_DIM).type(dtype)
    model.eval()
    model.load_state_dict(shared_model.state_dict())

    log_path_ini = LOG_FILE_VALID
    all_cooked_time, all_cooked_bw, all_file_names = load_trace.load_trace(TEST_TRACES_VALID)
    env = env_valid.Environment(all_cooked_time=all_cooked_time,
                                all_cooked_bw=all_cooked_bw,
                                set_idx=None)
    evaluation(model, log_path_ini, env, all_file_names, False)

    rewards = []
    test_log_files = os.listdir(TEST_LOG_FOLDER)
    for test_log_file in test_log_files:
        reward = []
        with open(TEST_LOG_FOLDER + test_log_file, 'rb') as f:
            for line in f:
                parse = line.split()
                try:
                    reward.append(float(parse[-1]))
                except IndexError:
                    break
        rewards.append(np.sum(reward[1:]))

    rewards = np.array(rewards)

    rewards_min = np.min(rewards)
    rewards_5per = np.percentile(rewards, 5)
    rewards_mean = np.mean(rewards)
    rewards_median = np.percentile(rewards, 50)
    rewards_95per = np.percentile(rewards, 95)
    rewards_max = np.max(rewards)

    log_file.write(str(int(epoch)) + '\t' +
                   str(rewards_min) + '\t' +
                   str(rewards_5per) + '\t' +
                   str(rewards_mean) + '\t' +
                   str(rewards_median) + '\t' +
                   str(rewards_95per) + '\t' +
                   str(rewards_max) + '\n')
    log_file.flush()

    add_str = 'a2c'
    model_save_path = Log_path + "/%s_%s_%d.model" % (str('abr'), add_str, int(epoch))
    torch.save(shared_model.state_dict(), model_save_path)
    print(epoch, rewards_min, rewards_5per, rewards_mean, rewards_median, rewards_95per, rewards_max)
    print('finish valid')


def test(test_model, test_traces, log_file):
    model = Actor(A_DIM).type(dtype)
    model.eval()
    model.load_state_dict(torch.load(test_model, map_location=torch.device('cpu')))

    log_path_ini = log_file + 'log_test_a2c'
    all_cooked_time, all_cooked_bw, all_file_names = load_trace.load_trace(test_traces)
    env = env_test.Environment(all_cooked_time=all_cooked_time,
                               all_cooked_bw=all_cooked_bw,
                               set_idx=0)
    evaluation(model, log_path_ini, env, all_file_names, False)

    rewards = []
    test_log_files = os.listdir(TEST_LOG_FOLDER)
    for test_log_file in test_log_files:
        reward = []
        with open(TEST_LOG_FOLDER + test_log_file, 'rb') as f:
            for line in f:
                parse = line.split()
                try:
                    reward.append(float(parse[-1]))
                except IndexError:
                    break
        rewards.append(np.sum(reward[1:]))

    rewards = np.array(rewards)

    rewards_min = np.min(rewards)
    rewards_5per = np.percentile(rewards, 5)
    rewards_mean = np.mean(rewards)
    rewards_median = np.percentile(rewards, 50)
    rewards_95per = np.percentile(rewards, 95)
    rewards_max = np.max(rewards)
    print(1, rewards_min, rewards_5per, rewards_mean, rewards_median, rewards_95per, rewards_max)
